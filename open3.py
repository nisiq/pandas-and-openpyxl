from openpyxl import load_workbook

wbt = load_workbook('teste.xlsx')
wbp = load_workbook('professores.xlsx')

p = wbp['Sheet1']
t = wbt['P3']
doutores = []

t['A1'] = 'Nº'
t['B1'] = 'Nome'

for i in range(30):
    if p.cell(row=i+1, column=3).value == 'DOUTORADO':
        doutores.append(p.cell(row=i+1, column=2).value)
for i in range(len(doutores)):
    print(doutores[i])

t['A2']= 'Nª'
for i in range(len(doutores)):
    t[f'A{i+1}'] = i
t['B2']= 'Nome'
for i in range(len(doutores)):
    t[f'B{i+1}'] = doutores[i]

wbt.save('teste.xlsx')